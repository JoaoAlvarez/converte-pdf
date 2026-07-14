"""Conversões entre PDF e formatos do Office (Word, Excel, PowerPoint)."""
from __future__ import annotations

import os
import re

# Data no formato dd/mm ou dd/mm/aaaa, mesmo colada a letras (sem exigir espaço),
# mas sem capturar pedaços no meio de números maiores.
_DATE_RE = re.compile(r"(?<!\d)(\d{2}/\d{2}(?:/\d{2,4})?)(?!\d)")
# Valor monetário no padrão brasileiro: 1.234,56 (com ou sem sinal).
_MONEY_RE = re.compile(r"(?<![\d.,])(-?\d{1,3}(?:\.\d{3})*,\d{2})(?![\d.,])")


def pdf_to_word(pdf_path: str, output_docx: str, progress=None) -> str:
    """Converte um PDF em documento do Word (.docx).

    Usa a biblioteca pdf2docx (Python puro) — não precisa do Word instalado.
    """
    from pdf2docx import Converter

    if progress:
        progress(0, 1, "Convertendo PDF para Word (pode levar um tempo)...")

    cv = Converter(pdf_path)
    try:
        cv.convert(output_docx, start=0, end=None)
    finally:
        cv.close()

    if progress:
        progress(1, 1, "Concluído.")
    return output_docx


def _cluster_1d(values, gap):
    """Agrupa valores próximos (dentro de ``gap``) e devolve o centro de cada grupo."""
    if not values:
        return []
    values = sorted(values)
    groups = [[values[0]]]
    for v in values[1:]:
        if v - groups[-1][-1] <= gap:
            groups[-1].append(v)
        else:
            groups.append([v])
    return [sum(g) / len(g) for g in groups]


def _groups_1d(values, gap):
    """Como _cluster_1d, mas devolve os grupos (para contar membros)."""
    if not values:
        return []
    values = sorted(values)
    groups = [[values[0]]]
    for v in values[1:]:
        if v - groups[-1][-1] <= gap:
            groups[-1].append(v)
        else:
            groups.append([v])
    return groups


def _detect_two_columns(words, page_width):
    """Detecta páginas com duas colunas de lançamentos lado a lado (faturas).

    Sinal seguro: datas SOLTAS (uma palavra inteira dd/mm) agrupadas em duas
    posições horizontais distintas e densas. Extratos de coluna única não
    disparam — no Itaú a data vem colada ao texto (não é palavra-data solta) e
    no Banco do Brasil todas as datas ficam na mesma posição.

    Retorna o x de corte, ou None.
    """
    date_x = [w["x0"] for w in words if _DATE_RE.fullmatch((w["text"] or "").strip())]
    if len(date_x) < 6:
        return None
    groups = [g for g in _groups_1d(date_x, gap=0.05 * page_width) if len(g) >= 3]
    if len(groups) < 2:
        return None
    groups.sort(key=len, reverse=True)
    g1, g2 = groups[0], groups[1]
    # A coluna secundária precisa ser densa (≥30% da principal) para não
    # confundir datas soltas dentro de uma descrição com uma segunda coluna.
    if len(g2) < 0.30 * len(g1):
        return None
    c1, c2 = sum(g1) / len(g1), sum(g2) / len(g2)
    if abs(c2 - c1) < 0.20 * page_width:
        return None
    right_group = g1 if c1 > c2 else g2
    return min(right_group) - 0.03 * page_width  # corta um pouco antes das datas da direita


def _cluster_words(words, row_tol=3, col_gap=14):
    """Agrupa uma lista de palavras em linhas (por y) e colunas (por x)."""
    if not words:
        return []
    words = sorted(words, key=lambda w: (round(w["top"]), w["x0"]))
    rows, cur, cur_top = [], [], None
    for w in words:
        if cur_top is None or abs(w["top"] - cur_top) <= row_tol:
            cur.append(w)
            cur_top = w["top"] if cur_top is None else cur_top
        else:
            rows.append(cur)
            cur, cur_top = [w], w["top"]
    if cur:
        rows.append(cur)

    col_starts = _cluster_1d([w["x0"] for w in words], col_gap)
    if not col_starts:
        return []

    table = []
    for row in rows:
        cells = [""] * len(col_starts)
        for w in sorted(row, key=lambda w: w["x0"]):
            ci = min(range(len(col_starts)), key=lambda i: abs(col_starts[i] - w["x0"]))
            cells[ci] = (cells[ci] + " " + w["text"]).strip()
        if any(c.strip() for c in cells):
            table.append(cells)
    return _drop_empty_columns(table)


def _words_to_table(page):
    """Reconstrói a tabela de uma página pelas posições das palavras.

    Se a página tem duas colunas de lançamentos lado a lado (faturas de
    cartão), divide-a e empilha os dois lados; caso contrário, processa a
    página inteira como uma coluna só (extratos).
    """
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    if not words:
        return []

    split_x = _detect_two_columns(words, page.width)
    if split_x is not None:
        left = [w for w in words if w["x0"] < split_x]
        right = [w for w in words if w["x0"] >= split_x]
        return _normalize_wordcluster(_cluster_words(left)) + \
            _normalize_wordcluster(_cluster_words(right))
    return _normalize_wordcluster(_cluster_words(words))


def _drop_empty_columns(table, min_fill=0.05):
    """Remove colunas quase totalmente vazias (ruído do agrupamento por palavras)."""
    if not table:
        return table
    ncols = max(len(r) for r in table)
    keep = []
    for c in range(ncols):
        filled = sum(1 for r in table if c < len(r) and r[c].strip())
        if filled / len(table) >= min_fill:
            keep.append(c)
    if not keep:
        return table
    return [[(r[c] if c < len(r) else "") for c in keep] for r in table]


def _normalize_wordcluster(table):
    """Separa datas e valores colados ao texto em colunas próprias.

    Aplicado apenas ao caminho de agrupamento por palavras. Para cada linha:
      - a primeira data encontrada vira uma coluna "Data" no início;
      - um valor monetário colado a texto (numa célula que também tem letras)
        é movido para uma coluna "Valor" no final.
    As colunas auxiliares só são mantidas se ao menos uma linha as preencher.
    """
    if not table:
        return table

    date_found = value_found = False
    rows = []
    for row in table:
        date_val = ""
        extracted_value = ""
        new_cells = []
        for cell in row:
            text = cell
            if not date_val:
                m = _DATE_RE.search(text)
                if m:
                    date_val = m.group(1)
                    text = (text[: m.start()] + text[m.end():]).strip()
                    date_found = True
            # Só separa o valor quando ele está colado a texto (há letras na célula).
            if not extracted_value and any(ch.isalpha() for ch in text):
                mv = _MONEY_RE.search(text)
                if mv:
                    extracted_value = mv.group(1)
                    text = (text[: mv.start()] + text[mv.end():]).strip()
                    value_found = True
            new_cells.append(text)
        rows.append([date_val] + new_cells + [extracted_value])

    if not value_found:
        rows = [r[:-1] for r in rows]
    if not date_found:
        rows = [r[1:] for r in rows]
    return _drop_empty_columns(rows)


def _table_quality(table):
    """Retorna (linhas, colunas, células_preenchidas) para avaliar uma tabela."""
    if not table:
        return 0, 0, 0
    ncols = max(len(r) for r in table)
    filled = sum(1 for r in table for c in r if c and str(c).strip())
    return len(table), ncols, filled


def _is_useful(table):
    """Uma tabela vale a pena se tem ≥2 linhas, ≥2 colunas e conteúdo real."""
    rows, cols, filled = _table_quality(table)
    return rows >= 2 and cols >= 2 and filled >= 4


def _write_statement_excel(statement, output_xlsx: str) -> int:
    """Escreve uma fatura estruturada: uma aba 'Lançamentos' com todas as
    transações e uma aba 'Resumo'. Devolve a quantidade de transações."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"
    headers = ["Titular", "Final", "Data", "Descrição", "Parcela", "Cidade",
               "Valor (R$)", "Tipo"]
    ws.append(headers)

    count = 0
    for section in statement.sections:
        for t in section.transactions:
            parcela = ""
            if t.installment_current and t.installment_total:
                parcela = f"{t.installment_current:02d}/{t.installment_total:02d}"
            valor = t.amount_cents / 100.0
            if t.is_credit:
                valor = -valor
            ws.append([
                section.holder_name,
                section.last_four_digits,
                t.date,
                t.description,
                parcela,
                t.city or "",
                valor,
                "Crédito" if t.is_credit else "Débito",
            ])
            count += 1

    # Formata a coluna de valor como moeda.
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.number_format = "#,##0.00"

    resumo = wb.create_sheet("Resumo")
    resumo.append(["Banco/Marca", statement.brand])
    resumo.append(["Cartão", statement.card_number])
    resumo.append(["Vencimento", statement.due_date])
    resumo.append(["Mês de referência", statement.billing_month])
    resumo.append(["Total da fatura (R$)", statement.total_cents / 100.0])
    resumo.append(["Qtd. de lançamentos", count])
    resumo["B5"].number_format = "#,##0.00"

    wb.save(output_xlsx)
    return count


def _write_itau_extrato_excel(lancamentos, output_xlsx: str) -> int:
    """Escreve o extrato do Itaú em uma única aba, com colunas separadas e
    uma linha de totais no fim."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato"
    ws.append(["Data", "Descrição", "Entrada (R$)", "Saída (R$)", "Saldo (R$)"])

    tot_ent = tot_sai = 0.0
    for t in lancamentos:
        ws.append([t["data"], t["descricao"], t["entrada"], t["saida"], t["saldo"]])
        tot_ent += t["entrada"] or 0.0
        tot_sai += t["saida"] or 0.0

    ws.append([])
    ws.append(["", "TOTAIS", round(tot_ent, 2), round(tot_sai, 2), ""])

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00"

    wb.save(output_xlsx)
    return len(lancamentos)


def pdf_to_excel(pdf_path: str, output_xlsx: str, progress=None,
                 password: str | None = None) -> tuple[str, int]:
    """Converte um PDF em planilha do Excel (.xlsx).

    Estratégia em camadas:
      0. **Fatura de cartão reconhecida** (PicPay, Nubank, etc.): usa um parser
         específico do banco e gera uma planilha estruturada (data, descrição,
         parcela, cidade, valor), MAS só se a soma dos lançamentos reconcilia
         com o total da fatura — senão cai para o método genérico.
      1. Detecção por **grade desenhada** (pdfplumber). Bom p/ Banco do Brasil.
      2. Reconstrução pelas **posições das palavras** — extratos alinhados por
         espaçamento (Itaú) e faturas de duas colunas.
      3. Texto linha a linha, para nunca gerar planilha vazia.

    Retorna (caminho, quantidade_de_tabelas_encontradas).
    """
    import pdfplumber
    from openpyxl import Workbook

    # Camada 0: fatura de cartão estruturada (com trava de reconciliação).
    if progress:
        progress(0, 1, "Verificando se é uma fatura de cartão reconhecida...")
    try:
        from . import banks
        statement = banks.best_statement(pdf_path, password)
    except Exception:
        statement = None
    if statement is not None:
        n = _write_statement_excel(statement, output_xlsx)
        if progress:
            progress(1, 1, "Fatura reconhecida e estruturada.")
        return output_xlsx, n

    # Camada 0b: extrato de conta corrente do Itaú (planilha única, colunas
    # data/descrição/entrada/saída/saldo).
    if progress:
        progress(0, 1, "Verificando se é um extrato do Itaú...")
    try:
        from . import extrato_itau
        lancamentos, ok = extrato_itau.parse(pdf_path, password)
    except Exception:
        ok, lancamentos = False, []
    if ok:
        n = _write_itau_extrato_excel(lancamentos, output_xlsx)
        if progress:
            progress(1, 1, "Extrato do Itaú reconhecido e estruturado.")
        return output_xlsx, n

    wb = Workbook()
    wb.remove(wb.active)  # começa sem abas; criamos conforme o conteúdo
    tables_found = 0

    def _new_sheet(title: str):
        # Nomes de aba no Excel: máx. 31 caracteres e sem caracteres proibidos.
        safe = "".join(c for c in title if c not in r"[]:*?/\\")[:31] or "Planilha"
        return wb.create_sheet(title=safe)

    def _write(title, table):
        ws = _new_sheet(title)
        for row in table:
            ws.append([("" if cell is None else str(cell)) for cell in row])

    with pdfplumber.open(pdf_path, password=password) as pdf:
        total = len(pdf.pages)
        for page_no, page in enumerate(pdf.pages, start=1):
            if progress:
                progress(page_no, total, f"Analisando página {page_no} de {total}")

            # 1) Grade desenhada.
            grid_tables = [t for t in (page.extract_tables() or []) if _is_useful(t)]
            if grid_tables:
                for t_no, table in enumerate(grid_tables, start=1):
                    tables_found += 1
                    _write(f"Pag{page_no}_Tabela{t_no}", table)
                continue

            # 2) Agrupamento por posição das palavras (com divisão em duas
            #    colunas quando for fatura, e separação de datas/valores).
            wt = _words_to_table(page)
            if _is_useful(wt):
                tables_found += 1
                _write(f"Pag{page_no}", wt)
                continue

            # 3) Texto simples desta página (uma linha por linha do PDF).
            text = page.extract_text() or ""
            if text.strip():
                ws = _new_sheet(f"Pag{page_no}_texto")
                for line in text.splitlines():
                    ws.append([line])

    if not wb.sheetnames:
        _new_sheet("Vazio")

    if progress:
        progress(1, 1, "Salvando planilha...")
    wb.save(output_xlsx)
    return output_xlsx, tables_found


# Extensões do Office suportadas na conversão para PDF.
OFFICE_EXTENSIONS = {".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt", ".rtf", ".odt"}


def office_to_pdf(input_path: str, output_pdf: str, progress=None) -> str:
    """Converte Word/Excel/PowerPoint em PDF.

    Estratégia (apenas Windows, que é o alvo do executável):
      1. Usa o Microsoft Office via automação COM, se estiver instalado.
      2. Se não houver Office, tenta o LibreOffice em modo headless.

    Levanta RuntimeError com mensagem amigável se nenhum estiver disponível.
    """
    ext = os.path.splitext(input_path)[1].lower()
    if ext not in OFFICE_EXTENSIONS:
        raise ValueError(f"Formato não suportado: {ext}")

    if progress:
        progress(0, 1, "Preparando conversão...")

    # 1) Microsoft Office via COM (Windows)
    try:
        return _office_com_to_pdf(input_path, output_pdf, ext, progress)
    except _NoOfficeError:
        pass

    # 2) LibreOffice em modo headless
    if _libreoffice_to_pdf(input_path, output_pdf, progress):
        return output_pdf

    raise RuntimeError(
        "Para converter arquivos do Office (Word/Excel/PowerPoint) é preciso ter "
        "o Microsoft Office ou o LibreOffice instalado neste computador."
    )


class _NoOfficeError(Exception):
    """Uso interno: Microsoft Office não está disponível."""


def _office_com_to_pdf(input_path, output_pdf, ext, progress):
    """Automação COM do Microsoft Office. Só funciona no Windows."""
    try:
        import comtypes.client  # type: ignore
    except ImportError:
        raise _NoOfficeError()

    input_path = os.path.abspath(input_path)
    output_pdf = os.path.abspath(output_pdf)
    wdFormatPDF = 17
    xlTypePDF = 0
    ppSaveAsPDF = 32

    if progress:
        progress(0, 1, "Abrindo no Microsoft Office...")

    try:
        if ext in {".docx", ".doc", ".rtf", ".odt"}:
            app = comtypes.client.CreateObject("Word.Application")
            app.Visible = False
            doc = app.Documents.Open(input_path)
            try:
                doc.SaveAs(output_pdf, FileFormat=wdFormatPDF)
            finally:
                doc.Close()
                app.Quit()

        elif ext in {".xlsx", ".xls"}:
            app = comtypes.client.CreateObject("Excel.Application")
            app.Visible = False
            wb = app.Workbooks.Open(input_path)
            try:
                wb.ExportAsFixedFormat(xlTypePDF, output_pdf)
            finally:
                wb.Close(False)
                app.Quit()

        elif ext in {".pptx", ".ppt"}:
            app = comtypes.client.CreateObject("PowerPoint.Application")
            pres = app.Presentations.Open(input_path, WithWindow=False)
            try:
                pres.SaveAs(output_pdf, ppSaveAsPDF)
            finally:
                pres.Close()
                app.Quit()
        else:
            raise _NoOfficeError()
    except OSError:
        # CreateObject falha quando o Office não está instalado.
        raise _NoOfficeError()

    if progress:
        progress(1, 1, "Concluído.")
    return output_pdf


def _libreoffice_to_pdf(input_path, output_pdf, progress) -> bool:
    """Tenta o LibreOffice em modo headless. Retorna True se der certo."""
    import shutil
    import subprocess
    import tempfile

    soffice = shutil.which("soffice") or shutil.which("soffice.exe")
    if not soffice:
        # Caminhos comuns de instalação no Windows.
        for candidate in (
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ):
            if os.path.exists(candidate):
                soffice = candidate
                break
    if not soffice:
        return False

    if progress:
        progress(0, 1, "Convertendo com o LibreOffice...")

    with tempfile.TemporaryDirectory() as tmp:
        subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", tmp, input_path],
            check=True,
            timeout=600,
        )
        base = os.path.splitext(os.path.basename(input_path))[0]
        produced = os.path.join(tmp, base + ".pdf")
        if not os.path.exists(produced):
            return False
        import shutil as _sh
        _sh.move(produced, output_pdf)

    if progress:
        progress(1, 1, "Concluído.")
    return True
